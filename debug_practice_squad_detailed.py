import pandas as pd
from openpyxl import load_workbook

# Load workbook
wb = load_workbook("🏈 Dine Nasty Football 2025.xlsx", data_only=True)

# Define the sheets to process
sheets = ["Beamen Division", "Falco Division"]

# Define the columns and rows to read (1-based Excel coordinates)
columns = ['C', 'H', 'M']  # These correspond to each team's Player column in the 4-column groups
rows = list(range(20, 23)) + list(range(67, 70))  # Rows 20-22 and 67-69

# Convert Excel column letters to indices (A=1, B=2, ...)
def excel_column_to_index(column: str) -> int:
    """
    Convert Excel column letter to index (A=1, B=2, ..., Z=26, AA=27, ...).
    """
    result = 0
    for char in column:
        result = result * 26 + (ord(char) - ord('A') + 1)
    return result

col_indices = [excel_column_to_index(col) for col in columns]

# Process each sheet
for sheet_name in sheets:
    print(f"=== {sheet_name} ===")
    if sheet_name not in wb.sheetnames:
        continue
        
    ws = wb[sheet_name]
    
    # Process each cell in the specified positions
    print("Practice Squad Players:")
    for i, row in enumerate(rows):
        for j, col_idx in enumerate(col_indices):
            cell = ws.cell(row=row, column=col_idx)
            if cell.value is not None and str(cell.value).strip():
                print(f"  Row {row}, Column {columns[j]}: '{cell.value}'")
            else:
                print(f"  Row {row}, Column {columns[j]}: (empty)")
    print()