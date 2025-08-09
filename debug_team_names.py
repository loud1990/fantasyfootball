from openpyxl import load_workbook

# Load workbook
wb = load_workbook("🏈 Dine Nasty Football 2025.xlsx", data_only=True)

# Define the sheets to process
sheets = ["Beamen Division", "Falco Division"]

# Process each sheet
for sheet_name in sheets:
    print(f"=== {sheet_name} ===")
    if sheet_name not in wb.sheetnames:
        continue
        
    ws = wb[sheet_name]
    
    # Get team names for this sheet (in row 1, columns 0, 5, 10)
    team_names = []
    team_columns = [0, 5, 10]  # Columns where team names are located
    for col_idx in team_columns:
        cell = ws.cell(row=1, column=col_idx + 1)  # +1 because Excel is 1-based
        team_names.append(cell.value if cell.value is not None else "")
    
    print("Team names:", team_names)
    print()